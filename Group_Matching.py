# Group Matching Code 
# PURPOSE: To create pairs of IDs that are matched based on a numerical matching column and gender (if applicable)
# Created by Michael Khela on 05/22/2024

# Import necessary packages
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font
import os

# INSERT your file path to "Matching_Package" (don't forget / at the end)
root_filepath = r"/Users/michaelkhela/Desktop/Matching_Package/"

# INSERT the name your comparison file
filename = "matching_export.csv"

# INSERT the comparison_group for comparison against control_group. PLEASE SPLIT BY COMMAS (unless only one is used)
comparison_group = "ASD,DS,FXS".split(",") # Keep the .split(","), it is needed later on!

# INSERT the control_group for comparison against comparison_group 
control_group = "TD"

# INSERT the maximum numerical difference between matching_column for control and comparison groups 
max_match_diff = 2

# INSERT column headers 
id_column = "subject_id"
cohort_column = "redcap_event_name"
matching_column = "age_at_vist" # Can be any column, HOWEVER it must be a number (age, score, etc)
sex_column = "child_sex_confirm"  # If this is not provided, set it to None

# DO NOT EDIT BELOW ------------------------------------------------

# Import the file name to extract ending
file_ext = filename[-4:]

# Read the data from REDCap export into a DataFrame called matching_df
if file_ext == "xlsx":
    matching_df = pd.read_excel(root_filepath + 'Inputs/' + filename)
elif file_ext == ".csv":
    matching_df = pd.read_csv(root_filepath + 'Inputs/' + filename)

# Select columns based on provided names
columns_to_use = [id_column, cohort_column, matching_column]
if sex_column:
    columns_to_use.append(sex_column)

matching_df = matching_df[columns_to_use]

# Drop rows with NaN values in the selected columns
matching_df = matching_df.dropna()

# Rename columns for consistency
matching_df.columns = ['id', 'Cohort', 'Matching'] + (['Sex'] if sex_column else [])

# Define the matching function
def matching(matching_df, comparison_group_list, control_group, max_match_diff, sex_column=None):
    matched_pairs = []
    unmatched_control = []
    unmatched_comparison = []

    # Extract data for each group
    control_df = matching_df[matching_df['Cohort'].str.contains(control_group)]
    comparison_df = matching_df[matching_df['Cohort'].str.contains('|'.join(comparison_group_list))]

    # Sort DataFrames by age
    sorted_control = control_df.sort_values(by='Matching')
    sorted_comparison = comparison_df.sort_values(by='Matching')

    # Track matched IDs for comparison group and control group per cohort
    matched_comparison_ids = set()
    matched_control_ids_per_cohort = {cohort: set() for cohort in comparison_group}

    # Iterate over each individual in comparison group
    for j in range(len(sorted_comparison)):
        comparison_row = sorted_comparison.iloc[j]
        subject_id2, age2, cohort2 = comparison_row[['id', 'Matching', 'Cohort']]
        gender2 = comparison_row['Sex'] if sex_column else None

        # Flag to check if a match is found
        match_found = False
        
        # Iterate over each individual in control group
        for i in range(len(sorted_control)):
            control_row = sorted_control.iloc[i]
            subject_id1, age1 = control_row[['id', 'Matching']]
            gender1 = control_row['Sex'] if sex_column else None
            
            # Check for gender match if sex_column is provided
            if sex_column:
                if gender1 != gender2:
                    continue
            
            # Check if the age difference is within the threshold
            age_diff = abs(age1 - age2)
            if age_diff <= max_match_diff and subject_id2 not in matched_comparison_ids and subject_id1 not in matched_control_ids_per_cohort[cohort2]:
                age_difference = age1 - age2
                matched_pair = (subject_id1, age1) + ((gender1,) if sex_column else ()) + (control_row['Cohort'], subject_id2, age2) + ((gender2,) if sex_column else ()) + (cohort2,) + (age_difference,)
                matched_pairs.append(matched_pair)
                matched_comparison_ids.add(subject_id2)
                matched_control_ids_per_cohort[cohort2].add(subject_id1)
                match_found = True
                break  # Found a match, so break the inner loop
        
        # If no match is found, add to unmatched for comparison group
        if not match_found:
            unmatched_comparison.append((subject_id2, age2) + ((gender2,) if sex_column else ()) + (cohort2,))

    # Append unmatched individuals from control group to unmatched for control group
    for i in range(len(sorted_control)):
        control_row = sorted_control.iloc[i]
        subject_id1, age1 = control_row[['id', 'Matching']]
        gender1 = control_row['Sex'] if sex_column else None
        if all(subject_id1 not in matched_control_ids_per_cohort[cohort] for cohort in comparison_group):
            unmatched_control.append((subject_id1, age1) + ((gender1,) if sex_column else ()) + (control_row['Cohort'],))

    return matched_pairs, unmatched_control, unmatched_comparison

# Run the age matching
matched_pairs, unmatched_control, unmatched_comparison = matching(matching_df, comparison_group, control_group, max_match_diff, sex_column)

# Get current date and time for the filename and header
now = datetime.now()
date_time_str = now.strftime("%Y-%m-%d")

# Create a function to check and create unique file name
def get_unique_filename(base_path, base_filename, extension):
    counter = 1
    unique_filename = f"{base_filename}{extension}"
    while os.path.exists(os.path.join(base_path, unique_filename)):
        unique_filename = f"{base_filename}_{counter}{extension}"
        counter += 1
    return unique_filename

# Generate the initial file name and check for uniqueness
base_filename = f"Matching_Results_{date_time_str}"
extension = ".xlsx"
output_filename = get_unique_filename(os.path.join(root_filepath, 'Outputs'), base_filename, extension)
output_filepath = os.path.join(root_filepath, 'Outputs', output_filename)

# Create an Excel writer object
writer = pd.ExcelWriter(output_filepath, engine='openpyxl')

# Convert results to DataFrames
matched_pairs_columns = ['Control_ID', 'Control_Matching'] + (['Control_Sex'] if sex_column else []) + ['Control_Cohort', 'Comparison_ID', 'Comparison_Matching'] + (['Comparison_Sex'] if sex_column else []) + ['Comparison_Cohort', 'Matching_Difference']
matched_pairs_df = pd.DataFrame(matched_pairs, columns=matched_pairs_columns)

unmatched_control_columns = ['Control_ID', 'Control_Matching'] + (['Control_Sex'] if sex_column else []) + ['Control_Cohort']
unmatched_control_df = pd.DataFrame(unmatched_control, columns=unmatched_control_columns)

unmatched_comparison_columns = ['Comparison_ID', 'Comparison_Matching'] + (['Comparison_Sex'] if sex_column else []) + ['Comparison_Cohort']
unmatched_comparison_df = pd.DataFrame(unmatched_comparison, columns=unmatched_comparison_columns)

# Write the data to Excel
matched_pairs_df.to_excel(writer, sheet_name='Matched Pairs', index=False)
unmatched_control_df.to_excel(writer, sheet_name='Unmatched Control', index=False)
unmatched_comparison_df.to_excel(writer, sheet_name='Unmatched Comparison', index=False)

# Close the writer to save the file
writer.close()

# Load the workbook to add header information
wb = openpyxl.load_workbook(output_filepath)
ws = wb['Matched Pairs']

# Add header information with bold words
header_labels = ['Comparison Groups:', 'Control Group:', 'Max Matching Difference:']
header_values = [', '.join(comparison_group), control_group, f"{max_match_diff}"]

for i, (label, value) in enumerate(zip(header_labels, header_values)):
    ws.insert_rows(1)
    label_cell = ws.cell(row=1, column=1, value=label)
    value_cell = ws.cell(row=1, column=2, value=value)
    label_cell.font = Font(bold=True)

# Save the final workbook
wb.save(output_filepath)

# Output file path
print(f"\nMatched Report is Created!!")
print(f"Output File: {output_filepath}")